"""
Une el texto extraido de los PDF con el dataset Excel de normas,
agregando una columna 'texto_pdf' a cada fila correspondiente.

El match se hace por: Tipo + Número + Año

Requisitos:
    pip install openpyxl
"""

import argparse
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


JSON   = "leyes.json"
EXCEL  = "Normativa_nacional_víctimas_1869-2023.xlsx"
OUTPUT = "Normativa_nacional_con_texto.xlsx"

# Correcciones manuales de metadatos del JSON 
# Estos 5 archivos tienen nombres irregulares y el parser no pudo extraer
# tipo/número/año automáticamente. Se corrigen acá explícitamente.

FIXES = {
    "1888 - Ley 2372 .pdf":                      ("Ley",     "2372",  1888),
    "1895 Decreto s-n.pdf":                       ("Decreto", "s/n",   1895),
    "1915 - Ley 9688 Accidentes de trabajo.pdf":  ("Ley",     "9688",  1915),
    "2011 - Ley 26683 Código Penal.pdf":          ("Ley",     "26683", 2011),
    "2014 - Ley 27063 Código Procesal Penal.pdf": ("Ley",     "27063", 2014),
}

# Equivalencias de tipo 
# El Excel usa "Decreto Reglamentario" para algunos decretos que en el
# JSON están guardados como "Decreto" (porque el nombre del archivo no
# incluye esa distinción). Esta tabla permite matchear igual.

TIPO_EQUIVALENCIAS = {
    "Decreto": ["Decreto", "Decreto Reglamentario"],
}


def build_lookup(data: list) -> dict:
    """
    Construye un diccionario (tipo, numero, year) -> texto
    a partir de la lista de documentos del JSON.
    """
    lookup = {} 

    for d in data:

        if d["filename"] in FIXES:
            tipo, numero, year = FIXES[d["filename"]]
        else:
            tipo   = d.get("tipo")
            numero = str(d.get("numero") or "")
            year   = d.get("year")

        if not tipo or not numero or not year:
            print(f"  [ADVERTENCIA] Sin metadatos completos: {d['filename']} — no se podrá matchear")
            continue

        tipos_posibles = TIPO_EQUIVALENCIAS.get(tipo, [tipo])
        for t in tipos_posibles:
            key = (t, numero, int(year))
            lookup[key] = d["text"]

    return lookup


def main():
    parser = argparse.ArgumentParser(description="Une texto OCR del JSON con el dataset Excel")
    parser.add_argument("--json",   default=JSON,   help=f"Archivo JSON con textos OCR (default: {JSON})")
    parser.add_argument("--excel",  default=EXCEL,  help=f"Dataset Excel de normas (default: {EXCEL})")
    parser.add_argument("--output", default=OUTPUT, help=f"Archivo Excel de salida (default: {OUTPUT})")
    args = parser.parse_args()

    json_path   = Path(args.json)
    excel_path  = Path(args.excel)
    output_path = Path(args.output)

    # Verificar que los archivos existen
    if not json_path.exists():
        print(f"ERROR: No se encontró el JSON: {json_path.resolve()}")
        return
    if not excel_path.exists():
        print(f"ERROR: No se encontró el Excel: {excel_path.resolve()}")
        return

    print(f"JSON  : {json_path.resolve()}")
    print(f"Excel : {excel_path.resolve()}")
    print(f"Salida: {output_path.resolve()}")
    print()

    # Cargar JSON
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    print(f"Documentos en JSON: {len(data)}")

    # Construir lookup
    lookup = build_lookup(data)
    print(f"Entradas en lookup: {len(lookup)}")
    print()

    # Cargar Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    print(f"Filas en Excel (sin header): {ws.max_row - 1}")
    print()

    # Agregar columna 'texto_pdf' al final
    last_col = ws.max_column + 1
    header_cell = ws.cell(row=1, column=last_col, value="texto_pdf")
    header_cell.font = Font(bold=True, name="Arial")
    header_cell.fill = PatternFill("solid", fgColor="D9E1F2")

    # Columnas del Excel (0-indexed): Tipo=0, Número=1, Año=2
    matched   = 0
    unmatched = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        tipo   = row[0].value
        numero = str(row[1].value) if row[1].value is not None else ""
        year   = row[2].value

        key  = (tipo, numero, year)
        text = lookup.get(key, "")

        cell = ws.cell(row=row[0].row, column=last_col, value=text)
        cell.alignment = Alignment(wrap_text=False)

        if text:
            matched += 1
        else:
            unmatched.append({"fila": row[0].row, "tipo": tipo, "numero": numero, "year": year})

    # Guardar
    wb.save(output_path)

    # Reporte final
    total = ws.max_row - 1
    print("=" * 50)
    print(f"Matcheados : {matched}/{total}")
    print(f"Sin match  : {len(unmatched)}")
    print(f"Salida     : {output_path.resolve()}")

    if unmatched:
        print("\nFilas sin match (revisar manualmente):")
        for u in unmatched:
            print(f"  Fila {u['fila']}: tipo={u['tipo']} numero={u['numero']} year={u['year']}")


if __name__ == "__main__":
    main()
